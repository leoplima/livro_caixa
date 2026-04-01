from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# ============= DASHBOARD =============
ws_dashboard = wb.create_sheet("Dashboard", 0)
ws_dashboard.sheet_properties.tabColor = "1F4E78"

# Largura das colunas
ws_dashboard.column_dimensions['A'].width = 25
ws_dashboard.column_dimensions['B'].width = 20
ws_dashboard.column_dimensions['C'].width = 20
ws_dashboard.column_dimensions['D'].width = 20
ws_dashboard.column_dimensions['E'].width = 20

# Cores modernas
cor_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
cor_entrada = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
cor_saida = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
cor_saldo = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cor_fundo = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

fonte_titulo = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
fonte_subtitulo = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
fonte_valor = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
fonte_label = Font(name='Calibri', size=11, color="1F4E78")

borda = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Título
ws_dashboard['A1'] = "LIVRO CAIXA DA IGREJA"
ws_dashboard['A1'].font = fonte_titulo
ws_dashboard['A1'].fill = cor_titulo
ws_dashboard.merge_cells('A1:E1')
ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_dashboard.row_dimensions[1].height = 35

# Data atual
ws_dashboard['A2'] = f"Extrato de: {datetime.now().strftime('%d/%m/%Y')}"
ws_dashboard['A2'].font = Font(name='Calibri', size=10, italic=True)
ws_dashboard.merge_cells('A2:E2')

# Seção de indicadores principais
row = 4
ws_dashboard[f'A{row}'] = "RESUMO FINANCEIRO"
ws_dashboard[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
ws_dashboard[f'A{row}'].fill = cor_titulo
ws_dashboard.merge_cells(f'A{row}:E{row}')
ws_dashboard[f'A{row}'].alignment = Alignment(horizontal='center')
ws_dashboard.row_dimensions[row].height = 25

# Cards de resumo
row = 5
ws_dashboard.row_dimensions[row].height = 60

# Total de Receitas
ws_dashboard[f'A{row}'] = "TOTAL DE RECEITAS"
ws_dashboard[f'A{row}'].font = fonte_subtitulo
ws_dashboard[f'A{row}'].fill = cor_entrada
ws_dashboard[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
ws_dashboard[f'A{row+1}'] = "=SUM(Transações!C:C)"
ws_dashboard[f'A{row+1}'].font = fonte_valor
ws_dashboard[f'A{row+1}'].fill = cor_entrada
ws_dashboard[f'A{row+1}'].number_format = 'R$ #,##0.00'
ws_dashboard[f'A{row+1}'].alignment = Alignment(horizontal='center')

# Total de Despesas
ws_dashboard[f'C{row}'] = "TOTAL DE DESPESAS"
ws_dashboard[f'C{row}'].font = fonte_subtitulo
ws_dashboard[f'C{row}'].fill = cor_saida
ws_dashboard[f'C{row}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
ws_dashboard[f'C{row+1}'] = "=SUM(Transações!D:D)"
ws_dashboard[f'C{row+1}'].font = fonte_valor
ws_dashboard[f'C{row+1}'].fill = cor_saida
ws_dashboard[f'C{row+1}'].number_format = 'R$ #,##0.00'
ws_dashboard[f'C{row+1}'].alignment = Alignment(horizontal='center')

# Saldo
ws_dashboard[f'E{row}'] = "SALDO TOTAL"
ws_dashboard[f'E{row}'].font = fonte_subtitulo
ws_dashboard[f'E{row}'].fill = cor_saldo
ws_dashboard[f'E{row}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
ws_dashboard[f'E{row+1}'] = "=A6-C6"
ws_dashboard[f'E{row+1}'].font = fonte_valor
ws_dashboard[f'E{row+1}'].fill = cor_saldo
ws_dashboard[f'E{row+1}'].number_format = 'R$ #,##0.00'
ws_dashboard[f'E{row+1}'].alignment = Alignment(horizontal='center')

# Gráfico de Receitas por Categoria
row = 9
ws_dashboard[f'A{row}'] = "RECEITAS POR CATEGORIA"
ws_dashboard[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color="1F4E78")

# Dados para o gráfico
categorias_receita = [
    ("Dízimos", 5000),
    ("Ofertas", 3500),
    ("Eventos", 2000),
    ("Doações", 1500),
    ("Outras Receitas", 800)
]

row_data = row + 1
for cat, valor in categorias_receita:
    ws_dashboard[f'A{row_data}'] = cat
    ws_dashboard[f'B{row_data}'] = valor
    ws_dashboard[f'A{row_data}'].font = fonte_label
    ws_dashboard[f'B{row_data}'].number_format = 'R$ #,##0.00'
    row_data += 1

# Gráfico Pizza - Receitas
pie1 = PieChart()
pie1.title = "Distribuição de Receitas"
pie1.style = 10
labels = Reference(ws_dashboard, min_col=1, min_row=row+1, max_row=row+5)
data = Reference(ws_dashboard, min_col=2, min_row=row, max_row=row+5)
pie1.add_data(data, titles_from_data=True)
pie1.set_categories(labels)
pie1.height = 10
pie1.width = 14
ws_dashboard.add_chart(pie1, f"A{row+7}")

# Gráfico de Despesas por Categoria
col_esp = 5
row = 9
ws_dashboard[f'E{row}'] = "DESPESAS POR CATEGORIA"
ws_dashboard[f'E{row}'].font = Font(name='Calibri', size=11, bold=True, color="1F4E78")

categorias_despesa = [
    ("Salários e Encargos", 4000),
    ("Manutenção", 1500),
    ("Utilitários", 800),
    ("Materiais", 600),
    ("Outras Despesas", 500)
]

row_data = row + 1
for cat, valor in categorias_despesa:
    ws_dashboard[f'E{row_data}'] = cat
    ws_dashboard[f'F{row_data}'] = valor
    ws_dashboard[f'E{row_data}'].font = fonte_label
    ws_dashboard[f'F{row_data}'].number_format = 'R$ #,##0.00'
    row_data += 1

# Gráfico Pizza - Despesas
pie2 = PieChart()
pie2.title = "Distribuição de Despesas"
pie2.style = 11
labels2 = Reference(ws_dashboard, min_col=5, min_row=row+1, max_row=row+5)
data2 = Reference(ws_dashboard, min_col=6, min_row=row, max_row=row+5)
pie2.add_data(data2, titles_from_data=True)
pie2.set_categories(labels2)
pie2.height = 10
pie2.width = 14
ws_dashboard.add_chart(pie2, f"G{row+7}")

# ============= PLANILHA DE TRANSAÇÕES =============
ws_transacoes = wb.create_sheet("Transações", 1)
ws_transacoes.sheet_properties.tabColor = "70AD47"

# Configurar colunas
ws_transacoes.column_dimensions['A'].width = 15
ws_transacoes.column_dimensions['B'].width = 20
ws_transacoes.column_dimensions['C'].width = 15
ws_transacoes.column_dimensions['D'].width = 15
ws_transacoes.column_dimensions['E'].width = 25
ws_transacoes.column_dimensions['F'].width = 15

# Cabeçalhos
headers = ['Data', 'Descrição', 'Receita (R$)', 'Despesa (R$)', 'Categoria', 'Saldo (R$)']
for col, header in enumerate(headers, 1):
    cell = ws_transacoes.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = borda

ws_transacoes.row_dimensions[1].height = 25

# Adicionar dados de exemplo
dados_exemplo = [
    (datetime.now() - timedelta(days=10), "Dízimo - Semana", 15000, 0, "Dízimos"),
    (datetime.now() - timedelta(days=9), "Aluguel do Templo", 0, 5000, "Aluguel"),
    (datetime.now() - timedelta(days=8), "Oferta - Domingo", 8500, 0, "Ofertas"),
    (datetime.now() - timedelta(days=7), "Contas (água, luz, gás)", 0, 2500, "Utilitários"),
    (datetime.now() - timedelta(days=6), "Dízimo - Semana", 12000, 0, "Dízimos"),
    (datetime.now() - timedelta(days=5), "Manutenção do templo", 0, 3000, "Manutenção"),
    (datetime.now() - timedelta(days=4), "Oferta - Acampamento", 5000, 0, "Ofertas"),
    (datetime.now() - timedelta(days=3), "Materiais de limpeza", 0, 800, "Materiais"),
    (datetime.now() - timedelta(days=2), "Dízimo - Semana", 16000, 0, "Dízimos"),
    (datetime.now() - timedelta(days=1), "Doação - Membro", 6000, 0, "Doações"),
]

saldo_acumulado = 0
for row_num, (data, desc, receita, despesa, categoria) in enumerate(dados_exemplo, 2):
    saldo_acumulado += receita - despesa
    
    ws_transacoes[f'A{row_num}'] = data
    ws_transacoes[f'A{row_num}'].number_format = 'dd/mm/yyyy'
    
    ws_transacoes[f'B{row_num}'] = desc
    ws_transacoes[f'C{row_num}'] = receita if receita > 0 else ''
    ws_transacoes[f'C{row_num}'].number_format = 'R$ #,##0.00'
    ws_transacoes[f'C{row_num}'].fill = PatternFill(start_color="E2EFD9", end_color="E2EFD9", fill_type="solid")
    
    ws_transacoes[f'D{row_num}'] = despesa if despesa > 0 else ''
    ws_transacoes[f'D{row_num}'].number_format = 'R$ #,##0.00'
    ws_transacoes[f'D{row_num}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    ws_transacoes[f'E{row_num}'] = categoria
    
    ws_transacoes[f'F{row_num}'] = saldo_acumulado
    ws_transacoes[f'F{row_num}'].number_format = 'R$ #,##0.00'
    ws_transacoes[f'F{row_num}'].fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
    
    for col in range(1, 7):
        ws_transacoes.cell(row=row_num, column=col).border = borda
        ws_transacoes.cell(row=row_num, column=col).alignment = Alignment(horizontal='center' if col in [1, 6] else 'left')

# ============= PLANILHA DE RELATÓRIO MENSAL =============
ws_relatorio = wb.create_sheet("Relatório Mensal", 2)
ws_relatorio.sheet_properties.tabColor = "4472C4"

ws_relatorio.column_dimensions['A'].width = 25
ws_relatorio.column_dimensions['B'].width = 18
ws_relatorio.column_dimensions['C'].width = 18

# Título
ws_relatorio['A1'] = f"RELATÓRIO MENSAL - {datetime.now().strftime('%B/%Y')}"
ws_relatorio['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
ws_relatorio['A1'].fill = cor_titulo
ws_relatorio.merge_cells('A1:C1')
ws_relatorio['A1'].alignment = Alignment(horizontal='center')
ws_relatorio.row_dimensions[1].height = 25

# Resumo
ws_relatorio['A3'] = "DESCRIÇÃO"
ws_relatorio['B3'] = "VALOR (R$)"
ws_relatorio['C3'] = "% DO TOTAL"

for col in ['A', 'B', 'C']:
    ws_relatorio[f'{col}3'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    ws_relatorio[f'{col}3'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_relatorio[f'{col}3'].alignment = Alignment(horizontal='center')
    ws_relatorio[f'{col}3'].border = borda

dados_relatorio = [
    ("Total de Receitas", "=SUM(Transações!C:C)"),
    ("Total de Despesas", "=SUM(Transações!D:D)"),
    ("Saldo do Mês", "=B4-B5"),
]

total_receita = "=B4"
for row_num, (desc, formula) in enumerate(dados_relatorio, 4):
    ws_relatorio[f'A{row_num}'] = desc
    ws_relatorio[f'B{row_num}'] = formula
    ws_relatorio[f'B{row_num}'].number_format = 'R$ #,##0.00'
    ws_relatorio[f'C{row_num}'] = f"=B{row_num}/{total_receita}" if row_num == 4 else f"=B{row_num}/{total_receita}"
    ws_relatorio[f'C{row_num}'].number_format = '0.00%'
    
    if desc == "Saldo do Mês":
        ws_relatorio[f'A{row_num}'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws_relatorio[f'A{row_num}'].fill = cor_saldo
        ws_relatorio[f'B{row_num}'].fill = cor_saldo
        ws_relatorio[f'B{row_num}'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws_relatorio[f'C{row_num}'].fill = cor_saldo
    else:
        ws_relatorio[f'A{row_num}'].fill = cor_fundo
        ws_relatorio[f'B{row_num}'].fill = cor_fundo
        ws_relatorio[f'C{row_num}'].fill = cor_fundo
    
    for col in ['A', 'B', 'C']:
        ws_relatorio[f'{col}{row_num}'].border = borda
        ws_relatorio[f'{col}{row_num}'].alignment = Alignment(horizontal='center' if col != 'A' else 'left')

# Gráfico de Evolução Mensal
row = 8
ws_relatorio[f'A{row}'] = "PRINCIPAIS CATEGORIAS"
ws_relatorio[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color="1F4E78")

categorias_principais = [
    "Dízimos",
    "Ofertas",
    "Manutenção",
    "Salários"
]

row_cat = row + 1
for cat in categorias_principais:
    ws_relatorio[f'A{row_cat}'] = cat
    ws_relatorio[f'B{row_cat}'] = random.randint(1000, 20000)
    ws_relatorio[f'B{row_cat}'].number_format = 'R$ #,##0.00'
    row_cat += 1

# Gráfico de Barras
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "Principais Categorias do Mês"
bar_chart.style = 11
data_bar = Reference(ws_relatorio, min_col=2, min_row=row, max_row=row+4)
cats_bar = Reference(ws_relatorio, min_col=1, min_row=row+1, max_row=row+4)
bar_chart.add_data(data_bar, titles_from_data=True)
bar_chart.set_categories(cats_bar)
bar_chart.height = 12
bar_chart.width = 18
ws_relatorio.add_chart(bar_chart, f"A{row+6}")

# Salvar arquivo
caminho_arquivo = r"c:\Users\Lekavi\Desktop\ia_master\Livro_Caixa_Igreja.xlsx"
wb.save(caminho_arquivo)

print(f"✓ Arquivo criado com sucesso: {caminho_arquivo}")
print("\nPlanilhas criadas:")
print("  1. Dashboard - Com indicadores e gráficos principais")
print("  2. Transações - Registro detalhado de entradas e saídas")
print("  3. Relatório Mensal - Resumo mensal e análises")
print("\nFuncionalidades:")
print("  ✓ Cores modernas e tema profissional")
print("  ✓ Gráficos de pizza das categorias")
print("  ✓ Gráfico de barras do mês")
print("  ✓ Cálculos automáticos de saldo")
print("  ✓ Formatação de moeda em Real")
print("  ✓ Fórmulas dinâmicas")

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import base64
import uuid
import hashlib
from pathlib import Path
from db import Database
from auth import Auth
from config import COLORS, CATEGORIAS_RECEITA, CATEGORIAS_DESPESA, ACCESS_LEVELS

# Função para converter hex para RGB
def hex_to_rgb(hex_color):
    """Converte cor hex (#RRGGBB) para RGB decimal"""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

# Obter RGB da cor primária e secundária
primary_rgb = hex_to_rgb(COLORS['primary'])
secondary_rgb = hex_to_rgb(COLORS['secondary'])
primary_rgb_str = f"{primary_rgb[0]}, {primary_rgb[1]}, {primary_rgb[2]}"
secondary_rgb_str = f"{secondary_rgb[0]}, {secondary_rgb[1]}, {secondary_rgb[2]}"

# Configuração da página
st.set_page_config(
    page_title="Livro Caixa - Igreja",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Aplicar marca d'água GLOBALMENTE (antes de tudo mais)
def apply_watermark_global():
    """Aplica marca d'água em TODAS as páginas via CSS global"""
    if 'current_church' in st.session_state and st.session_state.current_church:
        church = st.session_state.current_church
        if church.get('marca_agua_path'):
            try:
                with open(church.get('marca_agua_path'), 'rb') as f:
                    img_data = base64.b64encode(f.read()).decode()
                    watermark_css = f"""
                    <style>
                        .stApp::before {{
                            content: '';
                            position: fixed;
                            top: 50%;
                            left: 50%;
                            width: 500px;
                            height: 500px;
                            background: url('data:image/png;base64,{img_data}') no-repeat center;
                            background-size: contain;
                            opacity: 0.12;
                            transform: translate(-50%, -50%);
                            z-index: 0;
                            pointer-events: none;
                        }}
                        
                        .main {{
                            position: relative;
                            z-index: 1;
                        }}
                    </style>
                    """
                    st.markdown(watermark_css, unsafe_allow_html=True)
            except Exception as e:
                print(f"Erro ao carregar marca d'água: {e}")

# Chamar logo no início
apply_watermark_global()

# CSS MODERNO COM ANIMAÇÕES E RESPONSIVIDADE
st.markdown(f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
        
        * {{
            font-family: 'Plus Jakarta Sans', sans-serif;
        }}
        
        :root {{
            --primary: {COLORS['primary']};
            --secondary: {COLORS['secondary']};
            --accent: {COLORS['accent']};
            --bg: {COLORS['background']};
            --surface: {COLORS['surface']};
            --text: {COLORS['text']};
        }}
        
        html, body {{
            background: linear-gradient(135deg, {COLORS['background']} 0%, #1a2847 100%);
            color: {COLORS['text']};
        }}
        
        .stApp {{
            background: linear-gradient(135deg, {COLORS['background']} 0%, #1a2847 100%);
        }}
        
        /* ANIMAÇÕES */
        @keyframes slideIn {{
            from {{
                opacity: 0;
                transform: translateY(20px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; }}
            to {{ opacity: 1; }}
        }}
        
        @keyframes slideInLeft {{
            from {{
                opacity: 0;
                transform: translateX(-30px);
            }}
            to {{
                opacity: 1;
                transform: translateX(0);
            }}
        }}
        
        @keyframes pulse {{
            0%, 100% {{ opacity: 1; }}
            50% {{ opacity: 0.8; }}
        }}
        
        /* ELEMENTOS */
        .main {{
            animation: slideIn 0.6s ease-out;
        }}
        
        .stButton > button {{
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
            border: none;
            border-radius: 12px;
            padding: 12px 24px;
            font-weight: 600;
            transition: all 0.3s cubic-bezier(0.4, 0.0, 0.2, 1);
            box-shadow: 0 8px 16px rgba({primary_rgb_str}, 0.3);
            cursor: pointer;
            width: 100%;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-size: 14px;
        }}
        
        .stButton > button:hover {{
            transform: translateY(-2px);
            box-shadow: 0 12px 24px rgba({primary_rgb_str}, 0.4);
        }}
        
        .stButton > button:active {{
            transform: translateY(0);
        }}
        
        /* CARD STYLES */
        .metric-card {{
            background: linear-gradient(135deg, {COLORS['surface']} 0%, {COLORS['surface_light']} 100%);
            padding: 24px;
            border-radius: 16px;
            border-left: 5px solid var(--primary);
            animation: slideIn 0.6s ease-out;
            transition: all 0.3s ease;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2);
        }}
        
        .metric-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.3);
            border-left-color: var(--accent);
        }}
        
        .stMetric {{
            background: transparent !important;
            padding: 0 !important;
        }}
        
        /* TABS */
        .stTabs [role="tablist"] {{
            gap: 8px;
            padding: 16px 0;
            border-bottom: 2px solid {COLORS['surface_light']};
        }}
        
        .stTabs [role="tab"] {{
            background: {COLORS['surface']};
            border-radius: 12px 12px 0 0;
            color: {COLORS['text_secondary']};
            padding: 12px 24px;
            font-weight: 600;
            border: none;
            transition: all 0.3s ease;
        }}
        
        .stTabs [role="tab"][aria-selected="true"] {{
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
        }}
        
        /* INPUT FIELDS */
        .stTextInput input, .stNumberInput input, .stSelectbox select, 
        .stDateInput input, .stTextArea textarea {{
            background: {COLORS['surface']} !important;
            color: {COLORS['text']} !important;
            border: 2px solid {COLORS['surface_light']} !important;
            border-radius: 10px !important;
            padding: 12px 16px !important;
            transition: all 0.3s ease !important;
        }}
        
        .stTextInput input:focus, .stNumberInput input:focus, 
        .stSelectbox select:focus, .stDateInput input:focus, 
        .stTextArea textarea:focus {{
            border-color: var(--primary) !important;
            box-shadow: 0 0 0 3px rgba({primary_rgb_str}, 0.1) !important;
        }}
        
        /* SELECTBOX */
        .stSelectbox svg {{
            filter: invert(1);
        }}
        
        /* SIDEBAR */
        .stSidebar {{
            background: {COLORS['surface']} !important;
            animation: slideInLeft 0.6s ease-out;
        }}
        
        /* HEADINGS */
        h1, h2, h3, h4, h5, h6 {{
            color: {COLORS['text']} !important;
            font-weight: 700 !important;
        }}
        
        h1 {{
            background: linear-gradient(135deg, {COLORS['primary']} 0%, {COLORS['secondary']} 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 24px !important;
        }}
        
        /* EXPANDER */
        .stExpander {{
            background: {COLORS['surface']} !important;
            border-radius: 10px !important;
            border: 1px solid {COLORS['surface_light']} !important;
        }}
        
        /* SUCCESS/ERROR BOXES */
        .success-box {{
            background: rgba(16, 185, 129, 0.15) !important;
            padding: 16px !important;
            border-radius: 10px !important;
            border-left: 4px solid {COLORS['success']} !important;
            animation: slideIn 0.4s ease-out !important;
        }}
        
        .error-box {{
            background: rgba(239, 68, 68, 0.15) !important;
            padding: 16px !important;
            border-radius: 10px !important;
            border-left: 4px solid {COLORS['danger']} !important;
            animation: slideIn 0.4s ease-out !important;
        }}
        
        .warning-box {{
            background: rgba(245, 158, 11, 0.15) !important;
            padding: 16px !important;
            border-radius: 10px !important;
            border-left: 4px solid {COLORS['warning']} !important;
            animation: slideIn 0.4s ease-out !important;
        }}
        
        /* TABLE STYLES */
        .stDataFrame {{
            width: 100% !important;
        }}
        
        /* DIVIDER */
        .divider {{
            height: 2px;
            background: linear-gradient(90deg, transparent, {COLORS['surface_light']}, transparent);
            margin: 20px 0;
        }}
        
        /* MOBILE RESPONSIVENESS */
        @media (max-width: 768px) {{
            .stContainer {{
                padding: 8px !important;
            }}
            
            .stButton > button {{
                padding: 10px 16px !important;
                font-size: 12px !important;
            }}
            
            h1 {{
                font-size: 24px !important;
            }}
            
            h2 {{
                font-size: 18px !important;
            }}
            
            .metric-card {{
                padding: 16px !important;
            }}
        }}
        
        /* GRADIENT TEXT */
        .gradient-text {{
            background: linear-gradient(135deg, {COLORS['primary']} 0%, {COLORS['secondary']} 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        
        /* LOADING ANIMATION */
        .spinner {{
            animation: spin 1s linear infinite;
        }}
        
        @keyframes spin {{
            from {{ transform: rotate(0deg); }}
            to {{ transform: rotate(360deg); }}
        }}
        
        /* BADGE STYLES */
        .badge {{
            display: inline-block;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            background: linear-gradient(135deg, var(--primary) 0%, var(--secondary) 100%);
            color: white;
        }}
    </style>
""", unsafe_allow_html=True)

# INICIALIZAR SESSION STATE
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user = None
    st.session_state.user_id = None
    st.session_state.current_church = None
    st.session_state.page = 'dashboard'

# INSTÂNCIAS
db = Database()
auth = Auth()

# ===== FUNÇÕES AUXILIARES =====
def logout():
    st.session_state.logged_in = False
    st.session_state.user = None
    st.session_state.user_id = None
    st.session_state.current_church = None
    st.rerun()

def gerar_relatorio_profissional(transacoes, church, tipo_relatorio="mensal", mes=None, ano=None):
    """Gera um relatório profissional em Excel com múltiplas abas, gráficos e resumos"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        # Converter transações para DataFrame
        df_original = pd.DataFrame(transacoes)
        if df_original.empty:
            return None
        
        # Converter data para datetime
        df_original['data'] = pd.to_datetime(df_original['data'])
        
        # Calcular saldo do período anterior (para contexto)
        saldo_anterior = 0
        data_inicio_periodo = None
        
        if tipo_relatorio == "mensal" and mes and ano:
            # Calcular saldo até o mês anterior
            data_inicio_mes_atual = pd.Timestamp(year=ano, month=mes, day=1)
            df_anterior = df_original[df_original['data'] < data_inicio_mes_atual]
            saldo_anterior = (df_anterior[df_anterior['tipo'] == 'receita']['valor'].sum() - 
                            df_anterior[df_anterior['tipo'] == 'despesa']['valor'].sum())
            
            # Filtrar apenas o mês atual
            df = df_original[(df_original['data'].dt.month == mes) & (df_original['data'].dt.year == ano)]
            periodo_texto = f"Mês de {['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes-1]} de {ano}"
            mes_anterior_num = mes - 1 if mes > 1 else 12
            ano_anterior = ano if mes > 1 else ano - 1
            mes_anterior_texto = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes_anterior_num-1]
            
        elif tipo_relatorio == "anual" and ano:
            # Calcular saldo até o ano anterior
            data_inicio_ano = pd.Timestamp(year=ano, month=1, day=1)
            df_anterior = df_original[df_original['data'] < data_inicio_ano]
            saldo_anterior = (df_anterior[df_anterior['tipo'] == 'receita']['valor'].sum() - 
                            df_anterior[df_anterior['tipo'] == 'despesa']['valor'].sum())
            
            # Filtrar apenas o ano atual
            df = df_original[df_original['data'].dt.year == ano]
            periodo_texto = f"Ano de {ano}"
            ano_anterior = ano - 1
            
        else:
            df = df_original
            periodo_texto = "Período Completo"
        
        # Calcular métricas do período atual
        total_receitas = df[df['tipo'] == 'receita']['valor'].sum()
        total_despesas = df[df['tipo'] == 'despesa']['valor'].sum()
        saldo_periodo = total_receitas - total_despesas
        saldo_atual = saldo_anterior + saldo_periodo
        
        # Usar buffer em memória
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # ===== ABA 1: RESUMO EXECUTIVO =====
            ws_resumo = writer.book.create_sheet('📊 Resumo Executivo', 0)
            
            # Cores
            cor_header = PatternFill(start_color=COLORS['primary'].lstrip('#'), end_color=COLORS['primary'].lstrip('#'), fill_type="solid")
            cor_receita = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cor_despesa = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            cor_saldo = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
            cor_saldo_anterior = PatternFill(start_color="94A3B8", end_color="94A3B8", fill_type="solid")
            
            font_header = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            font_title = Font(name='Calibri', size=18, bold=True, color=COLORS['primary'].lstrip('#'))
            font_label = Font(name='Calibri', size=11, bold=True)
            font_valor = Font(name='Calibri', size=14, bold=True)
            
            # Header
            ws_resumo.merge_cells('A1:E1')
            ws_resumo['A1'] = f"RELATÓRIO FINANCEIRO - {church['nome'].upper()}"
            ws_resumo['A1'].font = font_title
            ws_resumo['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumo.row_dimensions[1].height = 30
            
            ws_resumo.merge_cells('A2:E2')
            ws_resumo['A2'] = periodo_texto
            ws_resumo['A2'].font = Font(name='Calibri', size=12, color="666666", italic=True)
            ws_resumo['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Espaço
            ws_resumo.row_dimensions[3].height = 10
            
            # Seção: Resumo Financeiro com Saldo Anterior
            ws_resumo['A4'] = "RESUMO FINANCEIRO"
            ws_resumo['A4'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            ws_resumo['A4'].fill = cor_header
            ws_resumo.merge_cells('A4:B4')
            
            # Saldo Anterior
            if tipo_relatorio in ["mensal", "anual"] and saldo_anterior != 0:
                ws_resumo['A5'] = "💰 Saldo Período Anterior:"
                ws_resumo['A5'].font = font_label
                ws_resumo['B5'] = saldo_anterior
                ws_resumo['B5'].font = font_valor
                ws_resumo['B5'].number_format = 'R$ #,##0.00'
                ws_resumo['A5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_resumo['B5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                row_inicio = 6
            else:
                row_inicio = 5
            
            # Total Receitas
            ws_resumo[f'A{row_inicio}'] = "📥 Total de Receitas:"
            ws_resumo[f'A{row_inicio}'].font = font_label
            ws_resumo[f'B{row_inicio}'] = total_receitas
            ws_resumo[f'B{row_inicio}'].font = font_valor
            ws_resumo[f'B{row_inicio}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'A{row_inicio}'].fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
            ws_resumo[f'B{row_inicio}'].fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
            
            # Total Despesas
            ws_resumo[f'A{row_inicio+1}'] = "📤 Total de Despesas:"
            ws_resumo[f'A{row_inicio+1}'].font = font_label
            ws_resumo[f'B{row_inicio+1}'] = total_despesas
            ws_resumo[f'B{row_inicio+1}'].font = font_valor
            ws_resumo[f'B{row_inicio+1}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'A{row_inicio+1}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            ws_resumo[f'B{row_inicio+1}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Saldo do Período
            ws_resumo[f'A{row_inicio+2}'] = "💵 Saldo do Período:"
            ws_resumo[f'A{row_inicio+2}'].font = font_label
            ws_resumo[f'B{row_inicio+2}'] = saldo_periodo
            ws_resumo[f'B{row_inicio+2}'].font = font_valor
            ws_resumo[f'B{row_inicio+2}'].number_format = 'R$ #,##0.00'
            cor_periodo = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            ws_resumo[f'A{row_inicio+2}'].fill = cor_periodo
            ws_resumo[f'B{row_inicio+2}'].fill = cor_periodo
            
            # Saldo Atual (Final)
            ws_resumo[f'A{row_inicio+3}'] = "💰 SALDO ATUAL:"
            ws_resumo[f'A{row_inicio+3}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            ws_resumo[f'B{row_inicio+3}'] = saldo_atual
            ws_resumo[f'B{row_inicio+3}'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            ws_resumo[f'B{row_inicio+3}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'A{row_inicio+3}'].fill = cor_saldo
            ws_resumo[f'B{row_inicio+3}'].fill = cor_saldo
            
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 18
            
            # Seção: Receitas por Categoria
            row = row_inicio + 5
            ws_resumo[f'A{row}'] = "RECEITAS POR CATEGORIA"
            ws_resumo[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            ws_resumo[f'A{row}'].fill = cor_receita
            ws_resumo.merge_cells(f'A{row}:B{row}')
            
            receitas_cat = df[df['tipo'] == 'receita'].groupby('categoria')['valor'].sum().sort_values(ascending=False)
            row += 1
            for cat, valor in receitas_cat.items():
                ws_resumo[f'A{row}'] = f"  {cat}"
                ws_resumo[f'B{row}'] = valor
                ws_resumo[f'B{row}'].number_format = 'R$ #,##0.00'
                row += 1
            
            # Seção: Despesas por Categoria
            row += 2
            ws_resumo[f'A{row}'] = "DESPESAS POR CATEGORIA"
            ws_resumo[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            ws_resumo[f'A{row}'].fill = cor_despesa
            ws_resumo.merge_cells(f'A{row}:B{row}')
            
            despesas_cat = df[df['tipo'] == 'despesa'].groupby('categoria')['valor'].sum().sort_values(ascending=False)
            row += 1
            for cat, valor in despesas_cat.items():
                ws_resumo[f'A{row}'] = f"  {cat}"
                ws_resumo[f'B{row}'] = valor
                ws_resumo[f'B{row}'].number_format = 'R$ #,##0.00'
                row += 1
            
            # ===== ABA 2: DETALHES DAS TRANSAÇÕES =====
            df_export = df[['data', 'descricao', 'tipo', 'valor', 'categoria', 'notas']].copy()
            df_export['data'] = df_export['data'].dt.strftime('%d/%m/%Y')
            df_export = df_export.sort_values('data', ascending=False)
            
            df_export.to_excel(writer, sheet_name='📋 Transações', index=False)
            ws_trans = writer.sheets['📋 Transações']
            
            # Formatar cabeçalho
            for cell in ws_trans[1]:
                cell.font = font_header
                cell.fill = cor_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Formatar dados
            for row in ws_trans.iter_rows(min_row=2, max_row=ws_trans.max_row):
                for cell in row:
                    if cell.column == 3:  # Coluna tipo
                        cell.value = "📥 Receita" if cell.value == "receita" else "📤 Despesa"
                    elif cell.column == 4:  # Coluna valor
                        cell.number_format = 'R$ #,##0.00'
            
            # Ajustar largura
            ws_trans.column_dimensions['A'].width = 12
            ws_trans.column_dimensions['B'].width = 25
            ws_trans.column_dimensions['C'].width = 12
            ws_trans.column_dimensions['D'].width = 14
            ws_trans.column_dimensions['E'].width = 18
            ws_trans.column_dimensions['F'].width = 20
            
            # ===== ABA 3: ANÁLISE POR CATEGORIA =====
            ws_cat = writer.book.create_sheet('📊 Por Categoria', 2)
            
            # Receitas por categoria - simplificado
            receitas_df = df[df['tipo'] == 'receita'].groupby('categoria').size().reset_index(name='Quantidade')
            receitas_df['Valor'] = df[df['tipo'] == 'receita'].groupby('categoria')['valor'].sum().reset_index(drop=True)
            receitas_df = receitas_df[['categoria', 'Valor', 'Quantidade']].sort_values('Valor', ascending=False)
            
            # Cabeçalho
            ws_cat['A1'] = "RECEITAS POR CATEGORIA"
            ws_cat['A1'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            ws_cat['A1'].fill = cor_receita
            ws_cat.merge_cells('A1:C1')
            
            # Headers de coluna
            ws_cat['A2'] = "Categoria"
            ws_cat['B2'] = "Valor"
            ws_cat['C2'] = "Qtd"
            for col_letter in ['A', 'B', 'C']:
                ws_cat[f'{col_letter}2'].font = Font(bold=True)
                ws_cat[f'{col_letter}2'].fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
            
            # Dados
            for idx, row in receitas_df.iterrows():
                ws_cat[f'A{idx+3}'] = row['categoria']
                ws_cat[f'B{idx+3}'] = row['Valor']
                ws_cat[f'B{idx+3}'].number_format = 'R$ #,##0.00'
                ws_cat[f'C{idx+3}'] = row['Quantidade']
            
            # Despesas por categoria - simplificado
            despesas_df = df[df['tipo'] == 'despesa'].groupby('categoria').size().reset_index(name='Quantidade')
            despesas_df['Valor'] = df[df['tipo'] == 'despesa'].groupby('categoria')['valor'].sum().reset_index(drop=True)
            despesas_df = despesas_df[['categoria', 'Valor', 'Quantidade']].sort_values('Valor', ascending=False)
            
            start_row = len(receitas_df) + 6
            
            # Cabeçalho
            ws_cat[f'A{start_row}'] = "DESPESAS POR CATEGORIA"
            ws_cat[f'A{start_row}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            ws_cat[f'A{start_row}'].fill = cor_despesa
            ws_cat.merge_cells(f'A{start_row}:C{start_row}')
            
            # Headers de coluna
            ws_cat[f'A{start_row+1}'] = "Categoria"
            ws_cat[f'B{start_row+1}'] = "Valor"
            ws_cat[f'C{start_row+1}'] = "Qtd"
            for col_letter in ['A', 'B', 'C']:
                ws_cat[f'{col_letter}{start_row+1}'].font = Font(bold=True)
                ws_cat[f'{col_letter}{start_row+1}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Dados
            for idx, row in despesas_df.iterrows():
                ws_cat[f'A{start_row+2+idx}'] = row['categoria']
                ws_cat[f'B{start_row+2+idx}'] = row['Valor']
                ws_cat[f'B{start_row+2+idx}'].number_format = 'R$ #,##0.00'
                ws_cat[f'C{start_row+2+idx}'] = row['Quantidade']
            
            # Ajustar coluna width
            ws_cat.column_dimensions['A'].width = 20
            ws_cat.column_dimensions['B'].width = 15
            ws_cat.column_dimensions['C'].width = 10
            
            # ===== ABA 4: TIMELINE DIÁRIA/MENSAL =====
            ws_timeline = writer.book.create_sheet('📈 Timeline', 3)
            
            if tipo_relatorio == "mensal":
                # Timeline diária para relatório mensal
                df_timeline = df.groupby(df['data'].dt.date).agg({
                    'valor': 'sum',
                    'tipo': lambda x: (x == 'receita').sum()
                }).reset_index()
                df_timeline.columns = ['Data', 'Valor Total', 'Qtd Receitas']
                df_timeline['Data'] = df_timeline['Data'].astype(str)
                timeline_label = "EVOLUÇÃO - POR DIA"
            else:
                # Timeline mensal para relatório anual
                df_timeline = df.groupby(df['data'].dt.to_period('M')).agg({
                    'valor': 'sum',
                    'tipo': lambda x: (x == 'receita').sum()
                }).reset_index()
                df_timeline['data'] = df_timeline['data'].astype(str)
                df_timeline.columns = ['Período', 'Valor Total', 'Qtd Receitas']
                timeline_label = "EVOLUÇÃO - POR MÊS"
            
            # Cabeçalho
            ws_timeline['A1'] = timeline_label
            ws_timeline['A1'].font = font_header
            ws_timeline['A1'].fill = cor_header
            
            # Escrever dados
            if tipo_relatorio == "mensal":
                ws_timeline['A2'] = "Data"
                ws_timeline['B2'] = "Valor Total"
                ws_timeline['C2'] = "Receitas"
            else:
                ws_timeline['A2'] = "Período"
                ws_timeline['B2'] = "Valor Total"
                ws_timeline['C2'] = "Receitas"
            
            for col in ['A', 'B', 'C']:
                ws_timeline[f'{col}2'].font = Font(bold=True)
                ws_timeline[f'{col}2'].fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid")
            
            for idx, row in df_timeline.iterrows():
                ws_timeline[f'A{idx+3}'] = row[0] if tipo_relatorio == "mensal" else row['Período']
                ws_timeline[f'B{idx+3}'] = row[1]
                ws_timeline[f'B{idx+3}'].number_format = 'R$ #,##0.00'
                ws_timeline[f'C{idx+3}'] = row[2]
            
            ws_timeline.column_dimensions['A'].width = 15
            ws_timeline.column_dimensions['B'].width = 15
            ws_timeline.column_dimensions['C'].width = 12
            
        # Retornar bytes do buffer
        buffer.seek(0)
        return buffer.getvalue()
    
    except Exception as e:
        st.error(f"Erro ao gerar relatório: {str(e)}")
        return None

def exportar_transacoes_excel(transacoes, nome_arquivo="transacoes.xlsx"):
    """Exporta transações simples para Excel (compatibilidade)"""
    try:
        # Criar dataframe
        df = pd.DataFrame(transacoes)
        
        # Formatar colunas
        if 'data' in df.columns:
            df['data'] = pd.to_datetime(df['data']).dt.strftime('%d/%m/%Y')
        if 'valor' in df.columns:
            df['valor'] = df['valor'].apply(lambda x: f"R$ {x:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.'))
        
        # Reordenar colunas
        colunas_ordem = ['data', 'descricao', 'tipo', 'valor', 'categoria', 'notas']
        df = df[[col for col in colunas_ordem if col in df.columns]]
        
        # Renomear colunas para display
        df = df.rename(columns={
            'data': 'Data',
            'descricao': 'Descrição',
            'tipo': 'Tipo',
            'valor': 'Valor',
            'categoria': 'Categoria',
            'notas': 'Observação'
        })
        
        # Exportar para Excel em memória
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transações', index=False)
            
            # Ajustar largura das colunas
            worksheet = writer.sheets['Transações']
            for idx, col in enumerate(df.columns, 1):
                max_length = max(df[col].astype(str).str.len().max(), len(col))
                worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 50)
        
        # Ler arquivo e retornar bytes
        with open(nome_arquivo, 'rb') as f:
            return f.read()
    except Exception as e:
        st.error(f"Erro ao exportar: {str(e)}")
        return None

def has_permission(permissao, access_level=None):
    """Verifica se o usuário tem a permissão"""
    if not st.session_state.logged_in:
        return False
    
    if access_level is None:
        access_level = st.session_state.user['access_level']
    
    return permissao in ACCESS_LEVELS[access_level]['permissoes']

def formato_moeda(valor):
    """Formata valor como moeda"""
    return f"R$ {valor:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')

def get_watermark_html(church):
    """Marca d'água agora é aplicada globalmente via CSS - função mantida para compatibilidade"""
    # Marca d'água já é aplicada por apply_watermark_global()
    return ""

# ===== TELA DE LOGIN =====
def tela_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("""
            <div style='text-align: center; animation: slideIn 0.8s ease-out;'>
                <h1 style='font-size: 48px; margin-bottom: 10px;'>💰 Livro Caixa</h1>
                <p style='font-size: 18px; color: #94A3B8; margin-bottom: 40px;'>
                    Gestão Financeira Inteligente para Igrejas
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["🔐 Login", "📝 Novo Usuário"])
        
        # TAB LOGIN
        with tab1:
            st.markdown('<div style="animation: slideIn 0.6s ease-out;">', unsafe_allow_html=True)
            
            email = st.text_input("📧 Email", key="login_email", placeholder="seu@email.com")
            senha = st.text_input("🔑 Senha", type="password", key="login_senha", placeholder="••••••••")
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔓 Entrar", use_container_width=True):
                    if not email or not senha:
                        st.error("⚠️ Preencha todos os campos!")
                    else:
                        sucesso, usuario, mensagem = auth.autenticar(email, senha)
                        if sucesso:
                            st.session_state.logged_in = True
                            st.session_state.user = usuario
                            st.session_state.user_id = usuario['id']
                            
                            # Obter primeira igreja disponível
                            igrejas = db.get_igrejas_usuario(usuario['id'])
                            if igrejas:
                                st.session_state.current_church = igrejas[0]
                            
                            st.success("✅ Bem-vindo!")
                            st.balloons()
                            st.rerun()
                        else:
                            st.error(f"❌ {mensagem}")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # TAB NOVO USUÁRIO
        with tab2:
            st.info("📧 Você será convidado por um administrador para criar sua conta")

# ===== TELA DO DASHBOARD =====
def tela_dashboard():
    church = st.session_state.current_church
    
    if not church:
        st.error("Nenhuma igreja selecionada")
        return
    
    # SELETOR DE IGREJA
    igrejas = db.get_igrejas_usuario(st.session_state.user_id)
    
    if len(igrejas) > 1:
        col1, col2, col3 = st.columns([2, 2, 3])
        with col1:
            igreja_selecionada = st.selectbox(
                "🏛️ Selecione a Igreja",
                igrejas,
                format_func=lambda x: x['nome'],
                key="select_church"
            )
            if igreja_selecionada and igreja_selecionada['id'] != church['id']:
                st.session_state.current_church = igreja_selecionada
                st.rerun()
    
    # Marca d'água
    st.markdown(get_watermark_html(church), unsafe_allow_html=True)
    
    # HEADER
    col1, col2, col3 = st.columns([1, 4, 2])
    with col1:
        st.markdown(f"### 🏛️ {church['nome']}")
    with col3:
        if st.button("👤 Sair", key="btn_logout"):
            logout()
    
    st.divider()
    
    # TÍTULO
    st.markdown("<h2 style='text-align: center;'>📊 Dashboard Financeiro</h2>", unsafe_allow_html=True)
    
    # FILTROS
    col1, col2, col3 = st.columns(3)
    with col1:
        data_inicio = st.date_input("📅 Data Inicial", datetime.now() - timedelta(days=30), key="dash_start")
    with col2:
        data_fim = st.date_input("📅 Data Final", datetime.now(), key="dash_end")
    with col3:
        st.write("")
    
    # OBTER DADOS
    transacoes = db.listar_transacoes(
        church['id'],
        data_inicio=data_inicio.strftime('%Y-%m-%d'),
        data_fim=data_fim.strftime('%Y-%m-%d')
    )
    
    resumo = db.get_resumo_financeiro(church['id'])
    
    # MÉTRICAS PRINCIPAIS
    st.markdown("<h3 style='margin-top: 30px;'>📈 Resumo Financeiro</h3>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
            <div class="metric-card">
                <div style='font-size: 14px; color: #94A3B8; margin-bottom: 8px;'>💰 Receitas</div>
                <div style='font-size: 32px; font-weight: 700; color: #10B981;'>{formato_moeda(resumo['receitas'])}</div>
            </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
            <div class="metric-card">
                <div style='font-size: 14px; color: #94A3B8; margin-bottom: 8px;'>💸 Despesas</div>
                <div style='font-size: 32px; font-weight: 700; color: #EF4444;'>{formato_moeda(resumo['despesas'])}</div>
            </div>
        """, unsafe_allow_html=True)
    
    with col3:
        cor_saldo = "#10B981" if resumo['saldo'] >= 0 else "#EF4444"
        st.markdown(f"""
            <div class="metric-card" style='border-left-color: {cor_saldo};'>
                <div style='font-size: 14px; color: #94A3B8; margin-bottom: 8px;'>📊 Saldo</div>
                <div style='font-size: 32px; font-weight: 700; color: {cor_saldo};'>{formato_moeda(resumo['saldo'])}</div>
            </div>
        """, unsafe_allow_html=True)
    
    # GRÁFICOS
    st.markdown("<h3 style='margin-top: 30px;'>📉 Análises</h3>", unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    # Gráfico: Receitas por Categoria
    with col1:
        receitas_cat = db.get_transacoes_por_categoria(church['id'], 'receita')
        if receitas_cat:
            df_receitas = pd.DataFrame(receitas_cat)
            fig = px.pie(
                df_receitas,
                values='total',
                names='categoria',
                color_discrete_sequence=[COLORS['primary'], COLORS['secondary'], 
                                       COLORS['success'], COLORS['warning'], 
                                       COLORS['info'], COLORS['danger']],
                title="Distribuição de Receitas"
            )
            fig.update_layout(
                font=dict(color=COLORS['text'], family="Plus Jakarta Sans"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sem dados de receita")
    
    # Gráfico: Despesas por Categoria
    with col2:
        despesas_cat = db.get_transacoes_por_categoria(church['id'], 'despesa')
        if despesas_cat:
            df_despesas = pd.DataFrame(despesas_cat)
            fig = px.pie(
                df_despesas,
                values='total',
                names='categoria',
                color_discrete_sequence=[COLORS['danger'], COLORS['warning'], 
                                       COLORS['info'], COLORS['success'], 
                                       COLORS['primary'], COLORS['secondary']],
                title="Distribuição de Despesas"
            )
            fig.update_layout(
                font=dict(color=COLORS['text'], family="Plus Jakarta Sans"),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sem dados de despesa")
    
    # Gráfico: Evolução Temporal
    if transacoes:
        st.markdown("<h3>📈 Evolução do Saldo</h3>", unsafe_allow_html=True)
        df = pd.DataFrame(transacoes)
        df['data'] = pd.to_datetime(df['data'])
        df = df.sort_values('data')
        
        df['valor_liquido'] = df.apply(lambda x: x['valor'] if x['tipo'] == 'receita' else -x['valor'], axis=1)
        df['saldo_acumulado'] = df['valor_liquido'].cumsum()
        
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df['data'],
            y=df['saldo_acumulado'],
            mode='lines+markers',
            name='Saldo',
            line=dict(color=COLORS['primary'], width=4),
            marker=dict(size=10, color=COLORS['secondary']),
            fill='tozeroy',
            fillcolor=f'rgba(124, 58, 237, 0.2)'
        ))
        
        fig.update_layout(
            font=dict(color=COLORS['text'], family="Plus Jakarta Sans"),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            xaxis_title='Data',
            yaxis_title='Saldo (R$)',
            hovermode='x unified',
            height=400
        )
        st.plotly_chart(fig, use_container_width=True)

# ===== TELA DE TRANSAÇÕES =====
def tela_transacoes():
    church = st.session_state.current_church
    
    if not church:
        st.error("Nenhuma igreja selecionada")
        return
    
    st.markdown(f"### 🏛️ {church['nome']}")
    st.divider()
    st.markdown("<h2>💳 Transações</h2>", unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["📋 Ver Transações", "➕ Nova Transação"])
    
    # TAB: VER TRANSAÇÕES
    with tab1:
        col1, col2, col3 = st.columns(3)
        with col1:
            filtro_tipo = st.selectbox("Tipo", ["Todos", "receita", "despesa"], key="filtro_tipo")
        with col2:
            filtro_categoria = st.selectbox("Categoria", ["Todas"] + CATEGORIAS_RECEITA + CATEGORIAS_DESPESA, key="filtro_categoria")
        with col3:
            st.write("")
        
        transacoes = db.listar_transacoes(
            church['id'],
            filtro_tipo=None if filtro_tipo == "Todos" else filtro_tipo,
            filtro_categoria=None if filtro_categoria == "Todas" else filtro_categoria
        )
        
        if transacoes:
            df = pd.DataFrame(transacoes)
            
            df_exibicao = df[['data', 'descricao', 'tipo', 'valor', 'categoria', 'notas']].copy()
            df_exibicao['valor'] = df_exibicao['valor'].apply(formato_moeda)
            df_exibicao['tipo'] = df_exibicao['tipo'].str.replace('receita', '📥').str.replace('despesa', '📤')
            
            st.dataframe(df_exibicao, use_container_width=True)
            
            # ===== SEÇÃO DE RELATÓRIOS E EXPORTAÇÃO =====
            st.divider()
            st.markdown("<h3>📊 Gerar Relatório Profissional</h3>", unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                tipo_relatorio = st.radio(
                    "Tipo de Relatório",
                    ["Simples", "Mensal", "Anual"],
                    horizontal=True,
                    key="tipo_relatorio"
                )
            
            # Opções de período baseado no tipo
            if tipo_relatorio == "Mensal":
                col_mes, col_ano = st.columns(2)
                with col_mes:
                    mes = st.selectbox(
                        "Mês",
                        list(range(1, 13)),
                        format_func=lambda x: ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                                              'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][x-1],
                        key="mes_relatorio"
                    )
                
                with col_ano:
                    ano = st.selectbox(
                        "Ano",
                        range(2020, datetime.now().year + 1),
                        index=(datetime.now().year - 2020),
                        key="ano_relatorio"
                    )
                
                # Botões de ação para relatório mensal
                col_rel1, col_rel2, col_rel3 = st.columns(3)
                
                with col_rel1:
                    if st.button("📄 Gerar Relatório Profissional", use_container_width=True, key="btn_rel_mensal"):
                        # Obter todas as transações (sem filtro de tipo/categoria)
                        todas_transacoes = db.listar_transacoes(church['id'])
                        
                        relatorio_bytes = gerar_relatorio_profissional(
                            todas_transacoes,
                            church,
                            tipo_relatorio="mensal",
                            mes=mes,
                            ano=ano
                        )
                        
                        if relatorio_bytes:
                            nome_arquivo = f"Relatorio_{church['nome']}_{['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'][mes-1]}_{ano}.xlsx"
                            st.download_button(
                                label="✅ Baixar Relatório Profissional",
                                data=relatorio_bytes,
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_prof_mensal"
                            )
                
                with col_rel2:
                    if st.button("📊 Exportar Simples (Excel)", use_container_width=True, key="btn_exp_simples_mensal"):
                        nome_arquivo = f"transacoes_{church['nome']}_{['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'][mes-1]}_{ano}.xlsx"
                        excel_bytes = exportar_transacoes_excel(transacoes, nome_arquivo)
                        if excel_bytes:
                            st.download_button(
                                label="✅ Baixar Excel",
                                data=excel_bytes,
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_simples_mensal"
                            )
            
            elif tipo_relatorio == "Anual":
                col_ano_anual, _ = st.columns(2)
                
                with col_ano_anual:
                    ano_anual = st.selectbox(
                        "Ano",
                        range(2020, datetime.now().year + 1),
                        index=(datetime.now().year - 2020),
                        key="ano_relatorio_anual"
                    )
                
                # Botões de ação para relatório anual
                col_rel1, col_rel2, col_rel3 = st.columns(3)
                
                with col_rel1:
                    if st.button("📄 Gerar Relatório Profissional", use_container_width=True, key="btn_rel_anual"):
                        # Obter todas as transações
                        todas_transacoes = db.listar_transacoes(church['id'])
                        
                        relatorio_bytes = gerar_relatorio_profissional(
                            todas_transacoes,
                            church,
                            tipo_relatorio="anual",
                            ano=ano_anual
                        )
                        
                        if relatorio_bytes:
                            nome_arquivo = f"Relatorio_{church['nome']}_{ano_anual}.xlsx"
                            st.download_button(
                                label="✅ Baixar Relatório Profissional",
                                data=relatorio_bytes,
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_prof_anual"
                            )
                
                with col_rel2:
                    if st.button("📊 Exportar Simples (Excel)", use_container_width=True, key="btn_exp_simples_anual"):
                        nome_arquivo = f"transacoes_{church['nome']}_{ano_anual}.xlsx"
                        excel_bytes = exportar_transacoes_excel(transacoes, nome_arquivo)
                        if excel_bytes:
                            st.download_button(
                                label="✅ Baixar Excel",
                                data=excel_bytes,
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_simples_anual"
                            )
            
            else:  # Simples
                col_rel1, col_rel2, col_rel3 = st.columns(3)
                
                with col_rel1:
                    if st.button("📊 Exportar para Excel", use_container_width=True, key="btn_exp_simples"):
                        nome_arquivo = f"transacoes_{church['nome']}_{pd.Timestamp.now().strftime('%d_%m_%Y')}.xlsx"
                        excel_bytes = exportar_transacoes_excel(transacoes, nome_arquivo)
                        if excel_bytes:
                            st.download_button(
                                label="✅ Baixar Excel",
                                data=excel_bytes,
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_simples"
                            )
            
            if has_permission('edit') or has_permission('delete'):
                st.divider()
                st.markdown("<h3>⚙️ Gerenciar</h3>", unsafe_allow_html=True)
                
                transacao_id = st.selectbox(
                    "Selecione uma transação",
                    options=[t['id'] for t in transacoes],
                    format_func=lambda x: f"{[t for t in transacoes if t['id'] == x][0]['data']} - {[t for t in transacoes if t['id'] == x][0]['descricao']}",
                    key="transacao_select"
                )
                
                col1, col2 = st.columns(2)
                with col1:
                    if has_permission('edit') and st.button("✏️ Editar", use_container_width=True):
                        st.session_state.edit_transacao_id = transacao_id
                        st.session_state.edit_mode = True
                
                with col2:
                    if has_permission('delete') and st.button("🗑️ Deletar", use_container_width=True):
                        db.deletar_transacao(transacao_id, st.session_state.user_id, church['id'])
                        st.success("Transação deletada!")
                        st.rerun()
            
            # Formulário de edição
            if hasattr(st.session_state, 'edit_mode') and st.session_state.edit_mode:
                st.divider()
                st.markdown("<h3>✏️ Editar Transação</h3>", unsafe_allow_html=True)
                
                transacao = [t for t in transacoes if t['id'] == st.session_state.edit_transacao_id]
                if transacao:
                    t = transacao[0]
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        tipo_edit = st.radio("Tipo", ["receita", "despesa"], 
                                           index=0 if t['tipo'] == 'receita' else 1,
                                           format_func=lambda x: "📥 Receita" if x == "receita" else "📤 Despesa",
                                           key="tipo_editar")
                    
                    with col2:
                        data_edit = st.date_input("Data", value=datetime.strptime(t['data'], '%Y-%m-%d'), key="data_editar")
                    
                    descricao_edit = st.text_input("Descrição", value=t['descricao'], key="descricao_editar")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        valor_edit = st.number_input("Valor (R$)", value=t['valor'], step=0.01, key="valor_editar")
                    
                    with col2:
                        categorias = CATEGORIAS_RECEITA if tipo_edit == "receita" else CATEGORIAS_DESPESA
                        categoria_edit = st.selectbox("Categoria", categorias, 
                                                     index=categorias.index(t['categoria']) if t['categoria'] in categorias else 0,
                                                     key="categoria_editar")
                    
                    notas_edit = st.text_area("📝 Observação", value=t.get('notas', '') or '', key="notas_editar")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("💾 Salvar Alterações", use_container_width=True):
                            db.atualizar_transacao(
                                st.session_state.edit_transacao_id,
                                st.session_state.user_id,
                                church['id'],
                                data=data_edit.strftime('%Y-%m-%d'),
                                descricao=descricao_edit,
                                tipo=tipo_edit,
                                valor=valor_edit,
                                categoria=categoria_edit,
                                notas=notas_edit
                            )
                            st.success("✅ Transação atualizada!")
                            st.session_state.edit_mode = False
                            del st.session_state.edit_transacao_id
                            st.rerun()
                    
                    with col2:
                        if st.button("❌ Cancelar", use_container_width=True):
                            st.session_state.edit_mode = False
                            del st.session_state.edit_transacao_id
                            st.rerun()
        else:
            st.info("Nenhuma transação encontrada")
    
    # TAB: NOVA TRANSAÇÃO
    with tab2:
        if not has_permission('edit'):
            st.error("Você não tem permissão para criar transações")
        else:
            col1, col2 = st.columns(2)
            
            with col1:
                tipo = st.radio("Tipo", ["receita", "despesa"], format_func=lambda x: "📥 Receita" if x == "receita" else "📤 Despesa", key="tipo_criar")
            
            with col2:
                data = st.date_input("Data", datetime.now(), key="data_criar")
            
            descricao = st.text_input("Descrição", placeholder="Ex: Dízimo da Semana", key="descricao_criar")
            
            col1, col2 = st.columns(2)
            with col1:
                valor = st.number_input("Valor (R$)", min_value=0.0, step=0.01, key="valor_criar")
            
            with col2:
                categorias = CATEGORIAS_RECEITA if tipo == "receita" else CATEGORIAS_DESPESA
                categoria = st.selectbox("Categoria", categorias, key="categoria_criar")
            
            notas = st.text_area("📝 Observação (opcional)", key="notas_criar")
            
            if st.button("💾 Salvar Transação", use_container_width=True):
                if not descricao or valor <= 0:
                    st.error("Preencha todos os campos obrigatórios")
                else:
                    db.criar_transacao(
                        data=data.strftime('%Y-%m-%d'),
                        descricao=descricao,
                        tipo=tipo,
                        valor=valor,
                        categoria=categoria,
                        usuario_id=st.session_state.user_id,
                        igreja_id=church['id'],
                        notas=notas
                    )
                    st.success("✅ Transação criada com sucesso!")
                    st.balloons()
                    st.rerun()

# ===== TELA DE USUÁRIOS (ADMIN) =====
def tela_usuarios():
    church = st.session_state.current_church
    
    if not church:
        st.error("Nenhuma igreja selecionada")
        return
    
    st.markdown(f"### 🏛️ {church['nome']}")
    st.divider()
    
    if not has_permission('manage_users'):
        st.error("❌ Você não tem permissão para acessar esta página")
        return
    
    st.markdown("<h2>👥 Gerenciar Usuários</h2>", unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["👤 Usuários", "📧 Convidar"])
    
    with tab1:
        # Verificar se o usuário tem permissão para gerenciar usuários (apenas admin)
        if st.session_state.user['access_level'] != 'admin':
            st.error("❌ Apenas administradores podem gerenciar usuários")
            return
        
        usuarios = db.listar_usuarios(church['id'])
        
        if usuarios:
            # Criar dataframe com informações adicionais (igrejas que o usuário tem acesso)
            usuarios_com_igrejas = []
            for u in usuarios:
                igrejas_acesso = db.get_igrejas_usuario(u['id'])
                nomes_igrejas = ', '.join([i['nome'] for i in igrejas_acesso])
                usuarios_com_igrejas.append({
                    'email': u['email'],
                    'nome': u['nome'],
                    'acesso_a': nomes_igrejas,
                    'access_level': u['access_level'],
                    'ativo': '✅ Ativo' if u.get('ativo') else '❌ Inativo'
                })
            
            df = pd.DataFrame(usuarios_com_igrejas)
            st.dataframe(df, use_container_width=True)
            
            st.divider()
            st.markdown("<h3>⚙️ Gerenciar Usuários</h3>", unsafe_allow_html=True)
            
            usuario_id = st.selectbox(
                "Selecione um usuário",
                options=[u['id'] for u in usuarios],
                format_func=lambda x: f"{[u for u in usuarios if u['id'] == x][0]['email']}",
                key="usuario_select"
            )
            
            # Obter dados do usuário selecionado
            usuario_selecionado = [u for u in usuarios if u['id'] == usuario_id][0] if usuario_id else None
            
            if usuario_selecionado:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("🔐 Resetar Senha", use_container_width=True):
                        try:
                            sucesso, msg = auth.redefinir_senha(usuario_id, st.session_state.user_id)
                            if sucesso:
                                st.success(msg)
                            else:
                                st.error(msg)
                        except Exception as e:
                            st.error(f"Erro ao resetar senha: {str(e)}")
                
                with col2:
                    if st.button("✏️ Editar Usuário", use_container_width=True):
                        st.session_state.edit_user_mode = True
                        st.session_state.edit_user_id = usuario_id
                        st.rerun()
                
                with col3:
                    status_atual = "Ativo ✅" if usuario_selecionado.get('ativo') else "Inativo ❌"
                    text_botao = "❌ Desativar" if usuario_selecionado.get('ativo') else "✅ Ativar"
                    if st.button(text_botao, use_container_width=True):
                        novo_status = 0 if usuario_selecionado.get('ativo') else 1
                        db.atualizar_usuario(usuario_id, ativo=novo_status)
                        st.success(f"✅ Usuário {'ativado' if novo_status else 'desativado'} com sucesso!")
                        st.rerun()
            
            # Modo de edição
            if st.session_state.get('edit_user_mode', False) and st.session_state.get('edit_user_id'):
                st.divider()
                st.markdown("<h3>✏️ Editar Usuário e Igrejas</h3>", unsafe_allow_html=True)
                
                usuario = [u for u in usuarios if u['id'] == st.session_state.edit_user_id][0]
                
                col_info1, col_info2 = st.columns(2)
                
                with col_info1:
                    nome_edit = st.text_input("Nome", value=usuario['nome'], key="edit_user_nome")
                    email_edit = st.text_input("Email", value=usuario['email'], key="edit_user_email")
                
                with col_info2:
                    novo_nivel_edit = st.selectbox("Nível de Acesso", list(ACCESS_LEVELS.keys()), 
                                                  index=list(ACCESS_LEVELS.keys()).index(usuario['access_level']),
                                                  key="edit_user_nivel")
                    ativo_edit = st.checkbox("Usuário Ativo", value=usuario.get('ativo', True), key="edit_user_ativo")
                
                # Seção para gerenciar igrejas
                st.divider()
                st.markdown("<h4>🏛️ Gerenciar Igrejas</h4>", unsafe_allow_html=True)
                
                # Obter igrejas que o usuário já tem acesso
                igrejas_com_acesso = db.get_igrejas_usuario(st.session_state.edit_user_id)
                igrejas_sem_acesso = db.get_igrejas_usuario_nao_tem_acesso(st.session_state.edit_user_id)
                
                tab_acesso, tab_adicionar = st.tabs(["✅ Igrejas com Acesso", "➕ Adicionar Igreja"])
                
                with tab_acesso:
                    if igrejas_com_acesso:
                        col_acesso1, col_acesso2 = st.columns([3, 1])
                        
                        for idx, ig in enumerate(igrejas_com_acesso):
                            col_acesso1, col_acesso2 = st.columns([3, 1])
                            
                            with col_acesso1:
                                st.info(f"🏛️ {ig['nome']}")
                            
                            with col_acesso2:
                                if st.button("🗑️ Remover", key=f"remove_church_{ig['id']}", use_container_width=True):
                                    db.remover_usuario_igreja(st.session_state.edit_user_id, ig['id'])
                                    st.success(f"✅ Acesso a {ig['nome']} removido!")
                                    st.rerun()
                    else:
                        st.warning("Usuário não tem acesso a nenhuma igreja")
                
                with tab_adicionar:
                    if igrejas_sem_acesso:
                        st.markdown("Clique no botão para adicionar acesso:")
                        
                        for idx, ig in enumerate(igrejas_sem_acesso):
                            col_add1, col_add2 = st.columns([3, 1])
                            
                            with col_add1:
                                st.info(f"🏛️ {ig['nome']}")
                            
                            with col_add2:
                                if st.button("➕ Adicionar", key=f"add_church_{ig['id']}", use_container_width=True):
                                    db.adicionar_usuario_igreja(st.session_state.edit_user_id, ig['id'], novo_nivel_edit)
                                    st.success(f"✅ Acesso a {ig['nome']} adicionado!")
                                    st.rerun()
                    else:
                        st.success("✅ Usuário já tem acesso a todas as igrejas!")
                
                st.divider()
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("💾 Salvar Alterações", use_container_width=True):
                        db.atualizar_usuario(
                            st.session_state.edit_user_id,
                            nome=nome_edit,
                            email=email_edit,
                            access_level=novo_nivel_edit,
                            ativo=1 if ativo_edit else 0
                        )
                        st.success("✅ Usuário atualizado com sucesso!")
                        st.session_state.edit_user_mode = False
                        del st.session_state.edit_user_id
                        st.rerun()
                
                with col2:
                    if st.button("❌ Cancelar", use_container_width=True):
                        st.session_state.edit_user_mode = False
                        del st.session_state.edit_user_id
                        st.rerun()
        else:
            st.info("Nenhum usuário cadastrado nesta igreja")
    
    with tab2:
        st.markdown("<h3>Convidar Novo Usuário</h3>", unsafe_allow_html=True)
        
        email = st.text_input("Email", key="convite_email")
        nome = st.text_input("Nome Completo", key="convite_nome")
        access_level = st.selectbox("Nível", list(ACCESS_LEVELS.keys()), key="nivel_convite")
        
        if st.button("📧 Enviar Convite", use_container_width=True):
            if not email or not nome:
                st.error("Preencha todos os campos")
            else:
                try:
                    sucesso, msg = auth.convidar_usuario(email, nome, access_level)
                    if sucesso:
                        # Obter o usuário criado
                        usuario = db.get_usuario_por_email(email)
                        if usuario:
                            db.adicionar_usuario_igreja(
                                usuario['id'],
                                church['id'],
                                access_level
                            )
                            st.success(msg)
                            st.balloons()
                            st.rerun()
                        else:
                            st.warning("Usuário criado mas não encontrado no banco. Tente novamente.")
                    else:
                        st.error(msg)
                except Exception as e:
                    st.error(f"Erro ao convidar usuário: {str(e)}")

# ===== TELA DE PERFIL =====
def tela_perfil():
    church = st.session_state.current_church
    
    st.markdown(f"### 🏛️ {church['nome'] if church else 'Perfil'}")
    st.divider()
    st.markdown("<h2>⚙️ Meu Perfil</h2>", unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("<h3>👤 Informações</h3>", unsafe_allow_html=True)
        st.info(f"**Email:** {st.session_state.user['email']}")
        st.info(f"**Nome:** {st.session_state.user['nome']}")
        st.info(f"**Nível:** {ACCESS_LEVELS[st.session_state.user['access_level']]['nome']}")
    
    with col2:
        st.markdown("<h3>🔐 Segurança</h3>", unsafe_allow_html=True)
        
        senha_atual = st.text_input("Senha Atual", type="password")
        senha_nova = st.text_input("Nova Senha", type="password")
        senha_confirma = st.text_input("Confirmar", type="password")
        
        if st.button("🔑 Alterar Senha", use_container_width=True):
            if not all([senha_atual, senha_nova, senha_confirma]):
                st.error("Preencha todos os campos")
            elif senha_nova != senha_confirma:
                st.error("As senhas não correspondem")
            else:
                sucesso, msg = auth.alterar_senha(st.session_state.user_id, senha_atual, senha_nova)
                if sucesso:
                    st.success(msg)
                else:
                    st.error(msg)

# ===== TELA DE IGREJAS (ADMIN) =====
def tela_igrejas():
    if not has_permission('manage_users'):
        st.error("❌ Você não tem permissão para acessar esta página")
        return
    
    st.markdown("<h2>🏛️ Gerenciar Igrejas</h2>", unsafe_allow_html=True)
    st.divider()
    
    tab1, tab2, tab3 = st.tabs(["📋 Igrejas", "➕ Nova Igreja", "📂 Categorias"])
    
    with tab1:
        igrejas = db.listar_igrejas()
        
        if igrejas:
            # Mostrar tabela com colunas específicas
            df = pd.DataFrame(igrejas)
            st.dataframe(df[['nome', 'cnpj', 'endereco', 'telefone', 'email']], use_container_width=True)
            
            st.divider()
            st.markdown("<h3>⚙️ Gerenciar Igrejas</h3>", unsafe_allow_html=True)
            
            # Seletor de Igreja
            igreja_id = st.selectbox(
                "Selecione uma igreja",
                options=[i['id'] for i in igrejas],
                format_func=lambda x: f"{[i for i in igrejas if i['id'] == x][0]['nome']}",
                key="select_church_edit"
            )
            
            # Obter dados da Igreja selecionada
            igreja_selecionada = [i for i in igrejas if i['id'] == igreja_id][0] if igreja_id else None
            
            if igreja_selecionada:
                st.markdown(f"**Igreja:** {igreja_selecionada['nome']}")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("✏️ Editar", use_container_width=True, key="btn_editar_igreja"):
                        st.session_state.edit_church_mode = True
                        st.session_state.edit_church_id = igreja_id
                        st.rerun()
                
                with col2:
                    if st.button("🗑️ Deletar", use_container_width=True, key="btn_deletar_igreja"):
                        # Confirmação de deleção
                        st.session_state.confirm_delete_church = True
                        st.session_state.delete_church_id = igreja_id
                
                with col3:
                    st.write("")  # Espaçador
                
                # Confirmação de deleção
                if st.session_state.get('confirm_delete_church', False):
                    st.warning(f"⚠️ Tem certeza que quer deletar **{igreja_selecionada['nome']}**? Esta ação é irreversível!")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("✅ Sim, Deletar", use_container_width=True):
                            if db.deletar_igreja(st.session_state.delete_church_id):
                                st.success("✅ Igreja deletada com sucesso!")
                                st.session_state.confirm_delete_church = False
                                st.rerun()
                            else:
                                st.error("❌ Erro ao deletar igreja")
                    
                    with col2:
                        if st.button("❌ Cancelar", use_container_width=True):
                            st.session_state.confirm_delete_church = False
                            st.rerun()
            
            # Modo de edição
            if st.session_state.get('edit_church_mode', False) and st.session_state.get('edit_church_id'):
                st.divider()
                st.markdown("<h3>✏️ Editar Igreja</h3>", unsafe_allow_html=True)
                
                igreja = db.get_Igreja(st.session_state.edit_church_id)
                
                if igreja:
                    nome_edit = st.text_input("Nome da Igreja", value=igreja['nome'], key="edit_church_nome")
                    cnpj_edit = st.text_input("CNPJ", value=igreja['cnpj'] or '', key="edit_church_cnpj")
                    endereco_edit = st.text_input("Endereço", value=igreja['endereco'] or '', key="edit_church_endereco")
                    telefone_edit = st.text_input("Telefone", value=igreja['telefone'] or '', key="edit_church_telefone")
                    email_edit = st.text_input("Email", value=igreja['email'] or '', key="edit_church_email")
                    
                    # Seção de imagem
                    st.divider()
                    st.markdown("<h4>🖼️ Logo/Marca d'água</h4>", unsafe_allow_html=True)
                    
                    # Mostrar imagem atual se houver
                    if igreja['marca_agua_path']:
                        try:
                            st.info(f"📌 Imagem atual: {igreja['marca_agua_path']}")
                        except:
                            st.warning("Imagem anterior não encontrada")
                    else:
                        st.info("Nenhuma imagem cadastrada ainda")
                    
                    # Upload de nova imagem
                    st.markdown("**Alterar imagem:**")
                    uploaded_logo_edit = st.file_uploader(
                        "Selecione uma nova imagem (PNG, JPG, JPEG)",
                        type=['png', 'jpg', 'jpeg'],
                        key="edit_church_logo"
                    )
                    
                    if uploaded_logo_edit:
                        st.success(f"✅ Imagem selecionada: {uploaded_logo_edit.name}")
                        # Pré-visualizar
                        st.image(uploaded_logo_edit, caption="Prévia da nova imagem", use_column_width=True)
                    
                    # Botões de ação
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("💾 Salvar Alterações", use_container_width=True, key="btn_save_church_edit"):
                            logo_path = None
                            
                            # Se upload nova imagem
                            if uploaded_logo_edit:
                                # Usar nome único com timestamp/uuid para evitar conflitos
                                file_ext = Path(uploaded_logo_edit.name).suffix
                                unique_filename = f"church_{uuid.uuid4().hex[:8]}{file_ext}"
                                logo_path = f"uploads/{unique_filename}"
                                
                                Path("uploads").mkdir(exist_ok=True)
                                with open(logo_path, "wb") as f:
                                    f.write(uploaded_logo_edit.getbuffer())
                            
                            # Atualizar dados da Igreja
                            update_data = {
                                'nome': nome_edit,
                                'cnpj': cnpj_edit,
                                'endereco': endereco_edit,
                                'telefone': telefone_edit,
                                'email': email_edit
                            }
                            
                            # Se tem nova imagem, adiciona ao update
                            if logo_path:
                                update_data['marca_agua_path'] = logo_path
                            
                            db.atualizar_Igreja(st.session_state.edit_church_id, **update_data)
                            st.success("✅ Igreja atualizada com sucesso!")
                            st.session_state.edit_church_mode = False
                            del st.session_state.edit_church_id
                            st.rerun()
                    
                    with col2:
                        if st.button("❌ Cancelar", use_container_width=True, key="btn_cancel_church_edit"):
                            st.session_state.edit_church_mode = False
                            del st.session_state.edit_church_id
                            st.rerun()
                    
                    with col3:
                        if st.button("🗑️ Remover Imagem", use_container_width=True, key="btn_remove_church_img"):
                            db.atualizar_Igreja(st.session_state.edit_church_id, marca_agua_path=None)
                            st.success("✅ Imagem removida com sucesso!")
                            st.rerun()
        else:
            st.info("Nenhuma igreja cadastrada")
    
    with tab2:
        st.markdown("<h3>Criar Nova Igreja</h3>", unsafe_allow_html=True)
        
        nome = st.text_input("Nome da Igreja", placeholder="Ex: Igreja Evangélica Central")
        cnpj = st.text_input("CNPJ (opcional)")
        endereco = st.text_input("Endereço")
        telefone = st.text_input("Telefone")
        email = st.text_input("Email")
        
        uploaded_logo = st.file_uploader("Logo/Marca d'água", type=['png', 'jpg', 'jpeg'])
        
        if st.button("✅ Criar Igreja", use_container_width=True):
            if not nome:
                st.error("Preencha o nome da igreja")
            else:
                logo_path = None
                if uploaded_logo:
                    # Usar nome único com timestamp/uuid para evitar conflitos
                    file_ext = Path(uploaded_logo.name).suffix
                    unique_filename = f"church_{uuid.uuid4().hex[:8]}{file_ext}"
                    logo_path = f"uploads/{unique_filename}"
                    
                    # Criar diretório se não existir
                    Path("uploads").mkdir(exist_ok=True)
                    
                    # Salvar arquivo
                    with open(logo_path, "wb") as f:
                        f.write(uploaded_logo.getbuffer())
                
                igreja_id = db.criar_igreja(nome, cnpj, endereco, telefone, email)
                if igreja_id:
                    if logo_path:
                        db.atualizar_Igreja(igreja_id, marca_agua_path=logo_path)
                    
                    st.success("✅ Igreja criada com sucesso!")
                    st.balloons()
                    st.rerun()
                else:
                    st.error("Erro ao criar igreja")
    
    with tab3:
        st.markdown("<h3>📂 Gerenciar Categorias</h3>", unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📥 Receitas")
            st.info(f"Categorias atuais: {', '.join(CATEGORIAS_RECEITA)}")
            nova_categoria_receita = st.text_input("Nova Categoria de Receita", placeholder="Ex: Esmolas")
            if st.button("➕ Adicionar Receita", use_container_width=True):
                if nova_categoria_receita and nova_categoria_receita not in CATEGORIAS_RECEITA:
                    st.info("ℹ️ Para adicionar categorias permanentes, edite o arquivo config.py")
                    st.code(f"CATEGORIAS_RECEITA.append('{nova_categoria_receita}')", language="python")
                elif nova_categoria_receita in CATEGORIAS_RECEITA:
                    st.warning("⚠️ Esta categoria já existe")
                else:
                    st.error("❌ Digite o nome da categoria")
        
        with col2:
            st.markdown("#### 📤 Despesas")
            st.info(f"Categorias atuais: {', '.join(CATEGORIAS_DESPESA)}")
            nova_categoria_despesa = st.text_input("Nova Categoria de Despesa", placeholder="Ex: Limpeza")
            if st.button("➕ Adicionar Despesa", use_container_width=True):
                if nova_categoria_despesa and nova_categoria_despesa not in CATEGORIAS_DESPESA:
                    st.info("ℹ️ Para adicionar categorias permanentes, edite o arquivo config.py")
                    st.code(f"CATEGORIAS_DESPESA.append('{nova_categoria_despesa}')", language="python")
                elif nova_categoria_despesa in CATEGORIAS_DESPESA:
                    st.warning("⚠️ Esta categoria já existe")
                else:
                    st.error("❌ Digite o nome da categoria")
        
        st.divider()
        st.markdown("<h4>📝 Como Editar Categorias Permanentemente:</h4>", unsafe_allow_html=True)
        st.markdown("""
        1. Abra o arquivo `config.py`
        2. Procure por `CATEGORIAS_RECEITA` ou `CATEGORIAS_DESPESA`
        3. Adicione a nova categoria à lista:
        
        ```python
        CATEGORIAS_RECEITA = [
            "Dízimos",
            "Ofertas",
            "Sua Nova Categoria",  # ← Adicione aqui
            # ...
        ]
        ```
        
        4. Salve o arquivo
        5. Reinicie a aplicação
        """)

# ===== MENU PRINCIPAL =====
def main():
    if not st.session_state.logged_in:
        tela_login()
    else:
        # SIDEBAR
        with st.sidebar:
            st.markdown("""
                <div style='text-align: center; padding: 20px 0;'>
                    <h2 style='margin: 0;'>💰 Livro Caixa</h2>
                </div>
            """, unsafe_allow_html=True)
            
            if st.session_state.current_church:
                st.markdown(f"**🏛️ {st.session_state.current_church['nome']}**")
            
            st.markdown(f"👤 {st.session_state.user['nome']}")
            st.divider()
            
            st.markdown("**📍 Navegação**")
            
            if st.button("📊 Dashboard", use_container_width=True, key="nav_dashboard"):
                st.session_state.page = 'dashboard'
                st.rerun()
            
            if st.button("💳 Transações", use_container_width=True, key="nav_trans"):
                st.session_state.page = 'transacoes'
                st.rerun()
            
            if st.button("⚙️ Perfil", use_container_width=True, key="nav_perfil"):
                st.session_state.page = 'perfil'
                st.rerun()
            
            if has_permission('manage_users'):
                if st.button("👥 Usuários", use_container_width=True, key="nav_users"):
                    st.session_state.page = 'usuarios'
                    st.rerun()
                
                if st.button("🏛️ Igrejas", use_container_width=True, key="nav_churches"):
                    st.session_state.page = 'igrejas'
                    st.rerun()
            
            st.divider()
            
            if st.button("🚪 Sair", use_container_width=True):
                logout()
        
        # CONTEÚDO PRINCIPAL
        if st.session_state.page == 'dashboard':
            tela_dashboard()
        elif st.session_state.page == 'transacoes':
            tela_transacoes()
        elif st.session_state.page == 'usuarios':
            tela_usuarios()
        elif st.session_state.page == 'perfil':
            tela_perfil()
        elif st.session_state.page == 'igrejas':
            tela_igrejas()

if __name__ == "__main__":
    main()
